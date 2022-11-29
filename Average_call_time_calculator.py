import openpyxl
import time


def excel_prep():
    is_data = True
    row_count = 1
    while is_data:
        row_count += 1
        data = agents_sheet.cell(row=row_count, column=1).value
        if data is None:
            is_data = False
    sheet_data = []
    for i in range(1, row_count):
        agent_data = {'id': agents_sheet.cell(row=i, column=1).value, 'name': agents_sheet.cell(row=i, column=2).value}
        sheet_data.append(agent_data)
    return sheet_data


def name_replace():
    is_data_b = True
    row_count_b = 1
    while is_data_b:
        row_count_b += 1
        data = monthly_sheet.cell(row=row_count_b, column=1).value
        if data is None:
            is_data_b = False
    rows = monthly_sheet.iter_rows(min_row=2, max_row=row_count_b - 1, min_col=5, max_col=6)
    return rows


def stripper(string):
    while string[-1] != '\\':
        string = string.rstrip(string[-1])

    return string


def keep_alive():
    while True:
        time.sleep(5)


def banner():
    logo = """
          ___                       ___           ___           ___
         /  /\          ___        /  /\         /  /\         /__/\\
        /  /::\        /  /\      /  /:/        /  /::\       |  |::\\
       /  /:/\:\      /  /:/     /  /:/        /  /:/\:\      |  |:|:\\
      /  /:/~/::\    /  /:/     /  /:/  ___   /  /:/  \:\   __|__|:|\:\\
     /__/:/ /:/\:\  /  /::\    /__/:/  /  /\ /__/:/ \__\:\ /__/::::| \:\\
     \  \:\/:/__\/ /__/:/\:\   \  \:\ /  /:/ \  \:\ /  /:/ \  \:\~~\__\/
      \  \::/      \__\/  \:\   \  \:\  /:/   \  \:\  /:/   \  \:\\
       \  \:\           \  \:\   \  \:\/:/     \  \:\/:/     \  \:\\
        \  \:\           \  \:\   \  \::/       \  \::/       \  \:\\
         \__\/            \__\/    \__\/         \__\/         \__\/

    -----------------------POWERED BY 1514_DIGITAL-----------------------

    """
    print(logo)


banner()

agents_path = "/home/johnh/Projects/Python Projects/Call stats project/Agents.xlsx"
monthly_path = "/home/johnh/Projects/Python Projects/Call stats project/CallMonitor_September2022.xlsx"

agents_data = openpyxl.load_workbook(agents_path)  # load excel file
agents_sheet = agents_data.active  # import active sheet
monthly_data = openpyxl.load_workbook(monthly_path)  # load excel file
monthly_sheet = monthly_data.active  # import active sheet

# while True:
#     try:
#         agents_excel = input("Paste Agents data excel file path: ")  # ask for input
#         agents_path = agents_excel.replace('\\', '''\\\\''')  # replace '\' with '\\'
#         agents_path = agents_path.replace('\"', "")  # remove quotation marks
#         agents_path = agents_path.strip()
#         agents_data = openpyxl.load_workbook(agents_path)  # load excel file
#         agents_sheet = agents_data.active  # import active sheet
#     except:
#         print("INVALID INPUT: TRY AGAIN")
#         continue
#
#     try:
#         monthly_excel = input("Paste Monthly data excel file path: ")  # ask for input
#         monthly_path = monthly_excel.replace('\\', '''\\\\''')  # replace '\' with '\\'
#         monthly_path = monthly_path.replace('\"', "")  # remove quotation marks
#         monthly_path = monthly_path.strip()
#         monthly_data = openpyxl.load_workbook(monthly_path)  # load excel file
#         monthly_sheet = monthly_data.active  # import active sheet
#         # print(path)
#     except:
#         print("INVALID INPUT: TRY AGAIN")
#         while True:
#             try:
#                 monthly_excel = input("Paste Monthly data excel file path: ")  # ask for input
#                 monthly_path = monthly_excel.replace('\\', '''\\\\''')  # replace '\' with '\\'
#                 monthly_path = monthly_path.replace('\"', "")  # remove quotation marks
#                 monthly_path = monthly_path.strip()
#                 monthly_data = openpyxl.load_workbook(monthly_path)  # load excel file
#                 monthly_sheet = monthly_data.active
#             except:
#                 print("INVALID INPUT: TRY AGAIN")
#                 continue
#             break
#     break


agent_import = excel_prep()
monthly_data = name_replace()
name = []
duration = []

# a very unneccecary loop, it goes about through 900 elements of a list
# this version has given its important proccesing to a imperative form of programming
# a new version is needed
for a, b in monthly_data:
    id_number = a.value
    duration.append(b.value)
    for row in agent_import:
        agent_id = row['id']
        agent_name = row['name']
        if id_number == agent_id:
            id_number = agent_name
            name.append(id_number)
dictionary = {}
for i in name:
    sub_dictionary = {i: ''}
    dictionary.update(sub_dictionary)
n = 0
keys = list(dictionary.keys())
for i in name:
    sub_list = []
    while True:
        try:
            x = name.index(keys[n])
        except ValueError:
            break
        sub_list.append((duration[x]))
        name.pop(x)
        duration.pop(x)
    dictionary[keys[n]] = sub_list
    n += 1
new_name = []
new_average = []
total_calls = []
for key, value in dictionary.items():
    x = len(value)
    y = sum(value)
    try:
        z = (y / x) / 60
    except ZeroDivisionError:
        z = 0
    talk_time = round(z, 1)
    new_name.append(key)
    new_average.append(talk_time)
    total_calls.append(x)
is_data = True
row_count = 1
while is_data:
    row_count += 1
    data = agents_sheet.cell(row=row_count, column=1).value
    if data is None:
        is_data = False
agents_sheet['A1'] = "Agent Name"
agents_sheet['B1'] = "Average Talk Time"
agents_sheet['c1'] = "Total Calls"
c = 0

# this is the part where the calculated data is saved
# the last two lines in the final file are leftovers from the original file i.e. agents_sheet
# the lines were not overwritten as expected because, long story short, it fell short
# to prevent this, make a new file instead of rewriting and renaming the old one

for i in range(2, row_count):
    try:
        agents_sheet.cell(row=i, column=1).value = str(new_name[c])
        agents_sheet.cell(row=i, column=2).value = new_average[c]
        agents_sheet.cell(row=i, column=3).value = total_calls[c]
# this is the quick fix is could come up with
# the code is comment-less and imperative
# we will devise a better solution
    except IndexError:
        break
    c += 1

#  file saviours, dont bother looking in here
save_name = 'Average Talk Time.xlsx'
save_path = '/home/johnh/Projects/Python Projects/Call stats project/'  # stripper(agents_path)
agents_data.save(save_path + save_name)

time.sleep(3)
print("Finished calculating Agents Average Call Time")

# save_path = save_path.replace('''\\\\''', '\\')
print(f"File saved at {save_path} as {save_name}")

keep_alive()
